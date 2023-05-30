from fastapi import FastAPI, File, UploadFile, Request, Response
from fastapi.responses import HTMLResponse, StreamingResponse
from typing import List
import tempfile

from io import BytesIO
from tempfile import NamedTemporaryFile
import openpyxl
from copy import copy
import datetime
import os

async def create_upload_files(files: List[UploadFile]):

    result = []
    filetemplate = next(filter(lambda f: f.filename == 'vzor.xlsx', files), None)
    fileforname = next(filter(lambda f: f.filename != 'vzor.xlsx', files), None)
    resultfilename = fileforname.filename
    
    if filetemplate is None:
        filetemplate = files[0]
        # mame vzor a mame alespon dalsi dva soubory?
        multimode = len(files) > 2
    else:
        # nemame vzor, mame nejmene dva soubory?
        multimode = len(files) > 1
    if multimode:
        resultfilename = "VsechnyVykazy.xlsx"

    filetemplatecontent = await filetemplate.read()
    print('*'*30, flush=True)
    print(multimode, flush=True)
    print(resultfilename, flush=True)
    for file in files:
        if file.filename == "vzor.xlsx":
            continue
        if file == filetemplate:
            content = filetemplatecontent
        else:
            content = await file.read()
        memory = BytesIO(content)
        wb = openpyxl.load_workbook(filename=memory, read_only=True)
        ws = wb["DataCelyRok"]

        for index, row in enumerate(ws.rows):
            if index == 0:
                continue

            names = ["name", "month", "date", "desc", "hours"]
            newRow = dict(zip(names, map(lambda item: item.value, row)))
            if newRow["date"] is None:
                print("has no date", newRow)
            elif not isinstance(newRow["date"], datetime.datetime):
                print("date has bad type", newRow)
            else:
                result.append(newRow)


    memory = BytesIO(filetemplatecontent)
    resultFile = openpyxl.load_workbook(filename=memory)

    resultFileCelyRok = resultFile["DataCelyRok"]
    resultFileWs = resultFile["ProTisk"]

    prevName = None
    prevMonth = None
    rowIndex = 16
    for item in result:
        currentName = item["name"]
        currentMonth = item["date"].month
        if not multimode:
            # zpracovavame pouze vzor a jeden soubor, budeme generovat listy po mesicich
            if (currentName != prevName) or (currentMonth != prevMonth):
                # v datech se zmenilo jmeno nebo mesic, vytvorime novy list
                # print(currentName, currentMonth)
                currentWs = resultFile.copy_worksheet(resultFileWs)
                currentWs.title = f"{currentName}_{currentMonth}"
                rowIndex = 16

                print(currentName)

                names = currentName.split(" ")
                currentWs[f"B10"] = names[0]
                currentWs[f"B11"] = names[1]

                currentWs[f"C12"] = datetime.datetime(
                    year=item["date"].year, month=item["date"].month, day=1
                )
                if item["date"].month == 12:
                    currentWs[f"E12"] = datetime.datetime(
                        year=item["date"].year + 1, month=1, day=1
                    ) + datetime.timedelta(days=-1)
                else:
                    currentWs[f"E12"] = datetime.datetime(
                        year=item["date"].year, month=item["date"].month + 1, day=1
                    ) + datetime.timedelta(days=-1)

                prevName = currentName
                prevMonth = currentMonth

            # currentWs.insert_rows(rowIndex)
            currentWs[f"A{rowIndex}"] = item["date"]
            currentWs[f"B{rowIndex}"] = item["desc"]
            currentWs[f"F{rowIndex}"] = item["hours"]
            # for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            #    currentWs[f'{col}{rowIndex}']._style = copy(currentWs[f'{col}{rowIndex+1}']._style)
            rowIndex = rowIndex + 1

        # kopirovani raw dat pro kontingencni tabulku
        resultFileCelyRok.insert_rows(2)
        resultFileCelyRok["A2"] = currentName
        resultFileCelyRok["B2"] = item["date"].month
        resultFileCelyRok["C2"] = item["date"]
        resultFileCelyRok["D2"] = item["desc"]
        resultFileCelyRok["E2"] = item["hours"]
        for col in ["A", "B", "C", "D", "E", "F"]:
            resultFileCelyRok[f"{col}2"]._style = copy(
                resultFileCelyRok[f"{col}3"]._style
            )
        # resultFileCelyRok.append([currentName, item['month'], item['date'], item['desc'], item['hours']])

    print("saving data", flush=True)
    with NamedTemporaryFile(dir=".") as tmp:
        tempfilename = tmp.name
        print("temp file name", tmp.name, flush=True)
    resultFile.save(tempfilename)
    print("data saved", flush=True)
    with open(tempfilename, mode="rb") as tmp:
        tmp.seek(0)
        print("data rewind", flush=True)
        stream = tmp.read()
    os.remove(tempfilename)
    print("have stream 1", len(stream), flush=True)
    print("have stream 2", len(stream), flush=True)
    headers = {"Content-Disposition": f'attachment; filename="{resultfilename}"'}
    return Response(stream, media_type="application/vnd.ms-excel", headers=headers)


def SingleFile(source):
    # cteni dat ze zdroje
    result = []
    memory = BytesIO(source)
    wb = openpyxl.load_workbook(filename=memory, read_only=True)
    ws = wb["DataCelyRok"]

    for index, row in enumerate(ws.rows):
        if index == 0:
            continue

        names = ["name", "month", "date", "desc", "hours"]
        newRow = dict(zip(names, map(lambda item: item.value, row)))
        if newRow["date"] is None:
            print("has no date", newRow)
        elif not isinstance(newRow["date"], datetime.datetime):
            print("date has bad type", newRow)
        else:
            result.append(newRow)

    #zapis dat
    memory = BytesIO(source)
    resultFile = openpyxl.load_workbook(filename=memory)

    resultFileCelyRok = resultFile["DataCelyRok"]
    resultFileWs = resultFile["ProTisk"]

    prevName = None
    prevMonth = None
    rowIndex = 16
    for item in result:
        currentName = item["name"]
        currentMonth = item["date"].month
        if (currentName != prevName) or (currentMonth != prevMonth):
            # v datech se zmenilo jmeno nebo mesic, vytvorime novy list
            # print(currentName, currentMonth)
            currentWs = resultFile.copy_worksheet(resultFileWs)
            currentWs.title = f"{currentName}_{currentMonth}"
            rowIndex = 16

            print(currentName)

            names = currentName.split(" ")
            
            #predpoklada se, ze je vyplneno
            #currentWs[f"B10"] = names[0]
            #currentWs[f"B11"] = names[1]

            currentWs[f"C12"] = datetime.datetime(
                year=item["date"].year, month=item["date"].month, day=1
            )
            if item["date"].month == 12:
                currentWs[f"E12"] = datetime.datetime(
                    year=item["date"].year + 1, month=1, day=1
                ) + datetime.timedelta(days=-1)
            else:
                currentWs[f"E12"] = datetime.datetime(
                    year=item["date"].year, month=item["date"].month + 1, day=1
                ) + datetime.timedelta(days=-1)

            prevName = currentName
            prevMonth = currentMonth

        # currentWs.insert_rows(rowIndex)
        currentWs[f"A{rowIndex}"] = item["date"]
        currentWs[f"B{rowIndex}"] = item["desc"]
        currentWs[f"F{rowIndex}"] = item["hours"]
        # for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        #    currentWs[f'{col}{rowIndex}']._style = copy(currentWs[f'{col}{rowIndex+1}']._style)
        rowIndex = rowIndex + 1
    
    return resultFile

import os
from sqlalchemy import select

def ExportModels(sessionMaker, DBModels):
    """returns a dict of lists of dict
    it is a dict of tables (list) containing a rows (dict)
    DBModels defines a list of models to export
    """

    def ToDict(dbRow, cols):
        "Converts a row (sqlalchemy model) into dict"
        result = {}
        for col in cols:
            result[col] = getattr(dbRow, col)
        return result

    result = {}
    for tableName, DBModel in DBModels.items():  # iterate over all models
        
        cols = [col.name for col in DBModel.metadata.tables[tableName].columns]

        # query for all items in a table
        stm = select(DBModel)
        with sessionMaker() as session:
            dbRows = session.execute(stm)
            dbData = dbRows.scalars()

            # convert all rows into list of dicts and
            # insert it as a new key-value pair into result
            result[tableName] = [ToDict(row, cols) for row in dbData]

    
    import json
    with open("systemdata.json", "w") as outfile:
        json.dump(result, outfile, indent=4, default=json_serial)

    return result

import datetime
import decimal
import json

def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""

    if isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.isoformat()
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    
    raise TypeError ("Type %s not serializable" % type(obj))

async def getSystemDataJSON():
    from sqlalchemy import MetaData
    from sqlalchemy.ext.automap import automap_base
    from sqlalchemy import create_engine, inspect, select

    from nogql_api.DBDefinitions import ComposeConnectionString
    from sqlalchemy.orm import sessionmaker

    Base = automap_base()
    connectionstring = ComposeConnectionString()
    connectionstring = connectionstring.replace(
        "postgresql+asyncpg", "postgresql+psycopg2"
    )
    engine = create_engine(connectionstring)
    print("Extracting metadata ...")
    Base.prepare(
        engine,
        reflect=True,
        # classname_for_table=fromTableToModelName,
        # name_for_collection_relationship=fromTableToRelationNName,
        # name_for_scalar_relationship=fromTableToRelation1Name
    )

    print("Creating graph for UML ...")

    def getModels(SQLAlchemyBase=Base):
        baseClasses = SQLAlchemyBase.classes
        result = {}
        for item in dir(baseClasses):
            if item.startswith("_"):
                continue
            result[item] = getattr(baseClasses, item)
        return result
    
    models = getModels()
    Session = sessionmaker(engine)
    result = ExportModels(Session, models)
    
    # print(models)
    # print(result)
    return json.dumps(result, indent=4, default=json_serial)


async def create_upload_file_one(files: List[UploadFile]):
    filetemplate = files[0]
    filetemplatecontent = await filetemplate.read()   
    resultFile = SingleFile(filetemplatecontent)
    print("saving data", flush=True)
    with NamedTemporaryFile(dir=".") as tmp:
        tempfilename = tmp.name
        print("temp file name", tmp.name, flush=True)
    resultFile.save(tempfilename)
    print("data saved", flush=True)
    with open(tempfilename, mode="rb") as tmp:
        tmp.seek(0)
        print("data rewind", flush=True)
        stream = tmp.read()
    os.remove(tempfilename)
    print("have stream 1", len(stream), flush=True)
    
    headers = {"Content-Disposition": f'attachment; filename="{filetemplate.filename}"'}
    return Response(stream, media_type="application/vnd.ms-excel", headers=headers)    
    
async def exportSchema():
    from sqlalchemy_schemadisplay import create_uml_graph
    from sqlalchemy import MetaData
    from sqlalchemy.ext.automap import automap_base
    from sqlalchemy import create_engine

    from nogql_api.DBDefinitions import ComposeConnectionString

    Base = automap_base()
    connectionstring = ComposeConnectionString()
    connectionstring = connectionstring.replace(
        "postgresql+asyncpg", "postgresql+psycopg2"
    )
    engine = create_engine(connectionstring)

    print("Extracting metadata ...")
    Base.prepare(
        engine,
        reflect=True,
        # classname_for_table=fromTableToModelName,
        # name_for_collection_relationship=fromTableToRelationNName,
        # name_for_scalar_relationship=fromTableToRelation1Name
    )

    print("Creating graph for UML ...")

    def getModels(SQLAlchemyBase=Base):
        baseClasses = SQLAlchemyBase.classes
        result = []
        for item in dir(baseClasses):
            if item.startswith("_"):
                continue
            result.append(getattr(baseClasses, item))
        return result

    mappers = [cls.__mapper__ for cls in getModels(SQLAlchemyBase=Base)]
    graph = create_uml_graph(
        mappers,
        show_operations=False,  # not necessary in this case
        show_multiplicity_one=False,  # some people like to see the ones, some don't
    )
    print("Writing UML...")

    with NamedTemporaryFile() as tmp:
        # graph.write_png('/output/img/uml.png') # write out the file
        # graph.write_svg('/output/img/uml.svg') # write out the file
        # graph.write_svg('/output/img/uml.svg') # write out the file
        graph.write_svg(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        # headers = {
        #    'Content-Disposition': 'attachment; filename="uml.svg"'
        # }
        return Response(stream, media_type="image/svg+xml")  # , headers=headers)
